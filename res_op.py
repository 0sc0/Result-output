#自动化是实现共产主义的先决条件，越自动，越共产
#半自动是社会主义，全自动是共产主义！
#上面的说法太中二了，而且把共产主义庸俗化了
import tkinter as tk
import openpyxl as op
from tkinter.filedialog import askdirectory
import os
from openpyxl.styles.borders import Border, Side

root = tk.Tk()
root.title("整理输出结果-Group Zhang")

##################################################
#GLOBAL初始化
re_kind = 27
num_loadcase = 0
num_group = 0
pcl_file_location = tk.StringVar()
name_loadcase = []
name_group = []

# 各个group的应力，二维数组，
# 第一层为各个组，顺序同name_group
# 第二层各项分别为该组的：von Mises，X+-，Y+-，Z+-，shear，梁单元判别项（0为该组没有梁单元应力）,梁单元x应力+-
# 第三层为各个工况，顺序同name_loadcase
group_stress = []

##############################################
#响应函数
#get并显示目标文件夹
def pcl_location():
    global pcl_file_location
    global l_pcl_location

    pcl_file_location = tk.filedialog.askdirectory()
    #首先清除先前的location显示
    l_pcl_location.grid_forget()
    l_pcl_location = tk.Label(root, text=pcl_file_location)
    l_pcl_location.grid(row=1, column=3)

#创建提结果pcl程序
def create_pcl():
    global pcl_file_location
    global num_loadcase
    global num_group
    global name_loadcase
    global name_group
    global re_kind
    global xdb_num
    global stress_groupname

    i = 1
    while group_text.get("%d.0" % i, "%d.end" % i):
        name_group.append(group_text.get("%d.0" % i, "%d.end" % i))
        num_group += 1
        i+=1

    i = 1
    while loadcase_text.get("%d.0" % i, "%d.end" % i):
        name_loadcase.append(loadcase_text.get("%d.0" % i, "%d.end" % i))
        num_loadcase += 1
        i+=1

    #输出文件
    opr_pcl = open(pcl_file_location + "\opr.pcl", mode='w')
    opr_pcl.write(
        """FUNCTION a()
    string group[64](VIRTUAL), grp[64](1), grp1[64](1), group1[64](64), load_case[64](VIRTUAL)
    integer i, j
    string aa[64]
    
    sys_allocate_array(load_case, 1, %d)\n
    sys_allocate_array(group, 1, %d)\n""" % (num_loadcase, num_group)
    )

    for i in range(0,num_group):
        opr_pcl.write('\t\tgroup(%d)=  "GR:%s"\n' % (i+1, name_group[i]))

    opr_pcl.write('\n')

    for i in range(0,num_loadcase):
        opr_pcl.write('\t\tload_case(%d)=  "%s"\n' % (i + 1, name_loadcase[i]))

    opr_pcl.write('\n')
    opr_pcl.write('\t\tFOR(j=1 to %d)\n' % (num_loadcase))
    opr_pcl.write('\t\t\tFOR(i=1 to %d)\n' % (num_group))

    start_location = os.getcwd()
    if re_kind == 0:
        common = open(start_location + "\\DO NOT DELETE" + "\\op2_opr_common.txt")
        pcl_common = common.read()
        opr_pcl.write(pcl_common)
        opr_pcl.close()
    else:
        xdb_num = e_xdb_num.get()
        common = open(start_location + "\\DO NOT DELETE" + "\\xdb_opr_common.txt")
        pcl_common = common.read()
        opr_pcl.write(pcl_common)
        opr_pcl.write(xdb_num)
        opr_pcl.write(""":Static Subcase", @ \n""")

        common = open(start_location + "\\DO NOT DELETE" + "\\xdb_opr_common2.txt")
        pcl_common = common.read()
        opr_pcl.write(pcl_common)
        opr_pcl.write(xdb_num)
        opr_pcl.write(""":Static Subcase", @ \n""")

        common = open(start_location + "\\DO NOT DELETE" + "\\xdb_opr_common3.txt")
        pcl_common = common.read()
        opr_pcl.write(pcl_common)
        opr_pcl.close()

        del name_group[0:num_group]
        del name_loadcase[0:num_loadcase]
        num_group = 0
        num_loadcase = 0

#读取无格式结果文件
def read_rpt():
    global pcl_file_location
    global num_loadcase
    global num_group
    global name_loadcase
    global name_group
    global group_stress

    #验证group和loadcase是否在之前的操作中被清空，如果被清空则重新读取
    if num_group == 0:
        #读取group
        i = 1
        while group_text.get("%d.0" % i, "%d.end" % i):
            name_group.append(group_text.get("%d.0" % i, "%d.end" % i))
            num_group += 1
            i+=1
        #读取loadcase
        i = 1
        while loadcase_text.get("%d.0" % i, "%d.end" % i):
            name_loadcase.append(loadcase_text.get("%d.0" % i, "%d.end" % i))
            num_loadcase += 1
            i+=1

    #初始化group_stress，三维数组
    group_stress = [[[0 for i in range(num_loadcase)]for j in range(11)]for k in range(num_group)]
    # print(group_stress)
    #读取结果文件
    shell_text = open(pcl_file_location + "/shell", mode = 'r')
    read_temp = shell_text.readlines()
    # print(len(read_temp))
    i = 1
    while read_temp[i]:
        if read_temp[i].count('Load Case:'):
            #获取工况名
            loadcase_temp = read_temp[i][read_temp[i].find('Load Case:')+11:read_temp[i].find(',')]
            #获取分组名
            j = i
            while read_temp[j]:
                if read_temp[j].count('GR'):
                    group_temp = read_temp[j][read_temp[j].find('GR')+3:]
                    group_temp = group_temp.replace('\n','')
                    group_temp = group_temp.strip()
                    break
                else:
                    j += 1
            #获取von Mises max应力
            j = i
            while read_temp[j]:
                if read_temp[j].count('von Mises'):
                    von_temp = read_temp[j+2][read_temp[j].find('von Mises'):]
                    von_temp = von_temp.replace('\n','')
                    von_temp = von_temp.lstrip()
                    break
                else:
                    j += 1
            #获取x min应力
            j = i
            while read_temp[j]:
                if read_temp[j].count('X Component'):
                    x_min_temp = read_temp[j + 1][read_temp[j].find('X Component'):]
                    x_min_temp = x_min_temp.replace('\n','')
                    x_min_temp = x_min_temp.lstrip()
                    break
                else:
                    j += 1
            #获取x max应力
            j = i
            while read_temp[j]:
                if read_temp[j].count('X Component'):
                    x_max_temp = read_temp[j + 2][read_temp[j].find('X Component'):]
                    x_max_temp = x_max_temp.replace('\n','')
                    x_max_temp = x_max_temp.lstrip()
                    break
                else:
                    j += 1
            #获取y min应力
            j = i
            while read_temp[j]:
                if read_temp[j].count('Y Component'):
                    y_min_temp = read_temp[j + 1][read_temp[j].find('Y Component'):]
                    y_min_temp = y_min_temp.replace('\n','')
                    y_min_temp = y_min_temp.lstrip()
                    break
                else:
                    j += 1
            #获取y max应力
            j = i
            while read_temp[j]:
                if read_temp[j].count('Y Component'):
                    y_max_temp = read_temp[j + 2][read_temp[j].find('Y Component'):]
                    y_max_temp = y_max_temp.replace('\n','')
                    y_max_temp = y_max_temp.lstrip()
                    break
                else:
                    j += 1
            #获取z min应力
            j = i
            while read_temp[j]:
                if read_temp[j].count('Z Component'):
                    z_min_temp = read_temp[j + 1][read_temp[j].find('Z Component'):]
                    z_min_temp = z_min_temp.replace('\n','')
                    z_min_temp = z_min_temp.lstrip()
                    break
                else:
                    j += 1
            #获取z max应力
            j = i
            while read_temp[j]:
                if read_temp[j].count('Z Component'):
                    z_max_temp = read_temp[j + 2][read_temp[j].find('Z Component'):]
                    z_max_temp = z_max_temp.replace('\n','')
                    z_max_temp = z_max_temp.lstrip()
                    break
                else:
                    j += 1
            #获取shear max应力
            j = i
            while read_temp[j]:
                if read_temp[j].count('Max Shear'):
                    shear_max_temp = read_temp[j + 2][read_temp[j].find('Max Shear'):]
                    shear_max_temp = shear_max_temp.replace('\n','')
                    shear_max_temp = shear_max_temp.lstrip()
                    break
                else:
                    j += 1

            #将temp值输入group_stress
            group_stress[name_group.index(group_temp)][0][name_loadcase.index(loadcase_temp)] = float(von_temp)
            group_stress[name_group.index(group_temp)][1][name_loadcase.index(loadcase_temp)] = float(x_min_temp)
            group_stress[name_group.index(group_temp)][2][name_loadcase.index(loadcase_temp)] = float(x_max_temp)
            group_stress[name_group.index(group_temp)][3][name_loadcase.index(loadcase_temp)] = float(y_min_temp)
            group_stress[name_group.index(group_temp)][4][name_loadcase.index(loadcase_temp)] = float(y_max_temp)
            group_stress[name_group.index(group_temp)][5][name_loadcase.index(loadcase_temp)] = float(z_min_temp)
            group_stress[name_group.index(group_temp)][6][name_loadcase.index(loadcase_temp)] = float(z_max_temp)
            group_stress[name_group.index(group_temp)][7][name_loadcase.index(loadcase_temp)] = float(shear_max_temp)
            # print(group_stress[name_group.index(group_temp)][0][name_loadcase.index(loadcase_temp)])
        i += 1
        if i == len(read_temp):
            break

    #读取梁单元结果
    beam_text = open(pcl_file_location + "/beam", mode='r')
    read_temp = beam_text.readlines()
    i = 1
    while read_temp[i]:
        if read_temp[i].count('Load Case:'):
            #获取工况名
            loadcase_temp = read_temp[i][read_temp[i].find('Load Case:')+11:read_temp[i].find(',')]
            #获取分组名
            j = i
            while read_temp[j]:
                if read_temp[j].count('GR'):
                    group_temp = read_temp[j][read_temp[j].find('GR')+3:]
                    group_temp = group_temp.replace('\n','')
                    group_temp = group_temp.strip()
                    break
                else:
                    j += 1
            #修改该组判别项为1
            group_stress[name_group.index(group_temp)][8][name_loadcase.index(loadcase_temp)] = 1
            # 获取X Component min应力
            j = i
            while read_temp[j]:
                if read_temp[j].count('X Component'):
                    beam_xmin_temp = read_temp[j + 1][read_temp[j].find('X Component'):]
                    beam_xmin_temp = beam_xmin_temp.replace('\n', '')
                    beam_xmin_temp = beam_xmin_temp.lstrip()
                    break
                else:
                    j += 1
            #获取X Component max应力
            j = i
            while read_temp[j]:
                if read_temp[j].count('X Component'):
                    beam_xmax_temp = read_temp[j+2][read_temp[j].find('X Component'):]
                    beam_xmax_temp = beam_xmax_temp.replace('\n','')
                    beam_xmax_temp = beam_xmax_temp.lstrip()
                    break
                else:
                    j += 1
            #输入group_stress组
            group_stress[name_group.index(group_temp)][9][name_loadcase.index(loadcase_temp)] = float(beam_xmin_temp)
            group_stress[name_group.index(group_temp)][10][name_loadcase.index(loadcase_temp)] = float(beam_xmax_temp)

        i += 1
        if i == len(read_temp):
            break

        # print(group_stress)
#创建Excel结果表格
def create_excel():
    global pcl_file_location
    global group_stress
    global num_group
    global name_group

    read_rpt()

    word = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    res_excel = op.Workbook()
    res_sheet = res_excel.active
    res_sheet.title = '板单元'
    #设定细边框格式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for i in range(0, num_group):
        #组名等标题
        res_sheet['A%d' % (i * (num_loadcase + 6) + 1)] = name_group[i]
        res_sheet['A%d' % (i * (num_loadcase + 6) + 2)] = '工况'
        res_sheet['B%d' % (i * (num_loadcase + 6) + 2)] = 'Max'
        res_sheet['C%d' % (i * (num_loadcase + 6) + 2)] = 'Min'
        res_sheet['D%d' % (i * (num_loadcase + 6) + 2)] = 'Max'
        res_sheet['E%d' % (i * (num_loadcase + 6) + 2)] = 'Min'
        res_sheet['F%d' % (i * (num_loadcase + 6) + 2)] = 'Max'
        res_sheet['G%d' % (i * (num_loadcase + 6) + 2)] = 'Min'
        res_sheet['H%d' % (i * (num_loadcase + 6) + 2)] = 'Max'
        res_sheet['I%d' % (i * (num_loadcase + 6) + 2)] = 'Max'
        res_sheet['A%d' % (i * (num_loadcase + 6) + 3 + num_loadcase)] = 'Max'
        res_sheet['A%d' % (i * (num_loadcase + 6) + 4 + num_loadcase)] = '许用值'
        res_sheet['A%d' % (i * (num_loadcase + 6) + 5 + num_loadcase)] = '%'
        for j in range(0, num_loadcase):
            res_sheet['A%d' % (i * (num_loadcase + 6) + 3 + j)] = name_loadcase[j]
        res_sheet['B%d' % (i * (num_loadcase + 6) + 1)] = 'Se'
        res_sheet.merge_cells('C%d:D%d' % (i * (num_loadcase + 6) + 1, i * (num_loadcase + 6) + 1))
        res_sheet['C%d' % (i * (num_loadcase + 6) + 1)] = 'Sx'
        res_sheet.merge_cells('E%d:F%d' % (i * (num_loadcase + 6) + 1, i * (num_loadcase + 6) + 1))
        res_sheet['E%d' % (i * (num_loadcase + 6) + 1)] = 'Sy'
        res_sheet.merge_cells('G%d:H%d' % (i * (num_loadcase + 6) + 1, i * (num_loadcase + 6) + 1))
        res_sheet['G%d' % (i * (num_loadcase + 6) + 1)] = 'Sz'
        res_sheet['I%d' % (i * (num_loadcase + 6) + 1)] = 'Tmax'
        #输出应力结果
        for column in word:
            for j in range(0, num_loadcase):
                res_sheet['%s%d' % (column, (i * (num_loadcase + 6) + 3 + j))] = group_stress[i][word.index(column)][j]
            if (column=='C')or(column=='E')or(column=='G'):
                res_sheet['%s%d' % (column, (i * (num_loadcase + 6) + 3 + num_loadcase))] = min(group_stress[i][word.index(column)])
            else:
                res_sheet['%s%d' % (column, (i * (num_loadcase + 6) + 3 + num_loadcase))] = max(group_stress[i][word.index(column)])
        res_sheet.merge_cells('C%d:D%d' % ((i * (num_loadcase + 6) + 4 + num_loadcase), (i * (num_loadcase + 6) + 4 + num_loadcase)))
        res_sheet.merge_cells('E%d:F%d' % ((i * (num_loadcase + 6) + 4 + num_loadcase), (i * (num_loadcase + 6) + 4 + num_loadcase)))
        res_sheet.merge_cells('G%d:H%d' % ((i * (num_loadcase + 6) + 4 + num_loadcase), (i * (num_loadcase + 6) + 4 + num_loadcase)))

        # #加边框
        # # res_sheet['A%d' % ((i-1)*(num_loadcase+6)+1):'I%d' % ((i+1)*(num_loadcase+6)-1)].border = thin_border
        # res_sheet.border = thin_border

    #输出梁单元结果
    res_beam = res_excel.create_sheet()
    res_beam.title = '梁单元'
    num_group_bar = 0
    # print(group_stress)
    for i in range(0, num_group):
        # print(i)
        # print(group_stress[i][0][0])
        if group_stress[i][8][0] == 1:
            num_group_bar += 1

    #每5个工况一个表格，j为表格数
    for j in range (0, int(num_group_bar/5)+1):
        # 固定标题
        res_beam['A%d' % (j * (num_loadcase + 6) + 1)] = '组'
        res_beam['A%d' % (j * (num_loadcase + 6) + 2)] = '工况'
        res_beam['B%d' % (j * (num_loadcase + 6) + 2)] = 'Min'
        res_beam['C%d' % (j * (num_loadcase + 6) + 2)] = 'Max'
        res_beam['D%d' % (j * (num_loadcase + 6) + 2)] = 'Min'
        res_beam['E%d' % (j * (num_loadcase + 6) + 2)] = 'Max'
        res_beam['F%d' % (j * (num_loadcase + 6) + 2)] = 'Min'
        res_beam['G%d' % (j * (num_loadcase + 6) + 2)] = 'Max'
        res_beam['H%d' % (j * (num_loadcase + 6) + 2)] = 'Min'
        res_beam['I%d' % (j * (num_loadcase + 6) + 2)] = 'Max'
        res_beam['J%d' % (j * (num_loadcase + 6) + 2)] = 'Min'
        res_beam['K%d' % (j * (num_loadcase + 6) + 2)] = 'Max'
        res_beam['A%d' % (j * (num_loadcase + 6) + 3 + num_loadcase)] = 'Min/Max'
        res_beam['A%d' % (j * (num_loadcase + 6) + 4 + num_loadcase)] = '许用值'
        res_beam['A%d' % (j * (num_loadcase + 6) + 5 + num_loadcase)] = '%'
        #工况名
        for i in range(0, num_loadcase):
            res_beam['A%d' % (j * (num_loadcase + 6) + 3 + i)] = name_loadcase[i]
        #合并单元格
        res_beam.merge_cells('B%d:C%d' % (j * (num_loadcase + 6) + 1, j * (num_loadcase + 6) + 1))
        res_beam.merge_cells('D%d:E%d' % (j * (num_loadcase + 6) + 1, j * (num_loadcase + 6) + 1))
        res_beam.merge_cells('F%d:G%d' % (j * (num_loadcase + 6) + 1, j * (num_loadcase + 6) + 1))
        res_beam.merge_cells('H%d:I%d' % (j * (num_loadcase + 6) + 1, j * (num_loadcase + 6) + 1))
        res_beam.merge_cells('J%d:K%d' % (j * (num_loadcase + 6) + 1, j * (num_loadcase + 6) + 1))
        res_beam.merge_cells(
            'B%d:C%d' % (j * (num_loadcase + 6) + 3 + num_loadcase, j * (num_loadcase + 6) + 3 + num_loadcase))
        res_beam.merge_cells(
            'D%d:E%d' % (j * (num_loadcase + 6) + 3 + num_loadcase, j * (num_loadcase + 6) + 3 + num_loadcase))
        res_beam.merge_cells(
            'F%d:G%d' % (j * (num_loadcase + 6) + 3 + num_loadcase, j * (num_loadcase + 6) + 3 + num_loadcase))
        res_beam.merge_cells(
            'H%d:I%d' % (j * (num_loadcase + 6) + 3 + num_loadcase, j * (num_loadcase + 6) + 3 + num_loadcase))
        res_beam.merge_cells(
            'J%d:K%d' % (j * (num_loadcase + 6) + 3 + num_loadcase, j * (num_loadcase + 6) + 3 + num_loadcase))

        res_beam.merge_cells('B%d:C%d' % (j * (num_loadcase + 6) + 4 + num_loadcase, j * (num_loadcase + 6) + 4 + num_loadcase))
        res_beam.merge_cells('D%d:E%d' % (j * (num_loadcase + 6) + 4 + num_loadcase, j * (num_loadcase + 6) + 4 + num_loadcase))
        res_beam.merge_cells('F%d:G%d' % (j * (num_loadcase + 6) + 4 + num_loadcase, j * (num_loadcase + 6) + 4 + num_loadcase))
        res_beam.merge_cells('H%d:I%d' % (j * (num_loadcase + 6) + 4 + num_loadcase, j * (num_loadcase + 6) + 4 + num_loadcase))
        res_beam.merge_cells('J%d:K%d' % (j * (num_loadcase + 6) + 4 + num_loadcase, j * (num_loadcase + 6) + 4 + num_loadcase))
        res_beam.merge_cells(
            'B%d:C%d' % (j * (num_loadcase + 6) + 5 + num_loadcase, j * (num_loadcase + 6) + 5 + num_loadcase))
        res_beam.merge_cells(
            'D%d:E%d' % (j * (num_loadcase + 6) + 5 + num_loadcase, j * (num_loadcase + 6) + 5 + num_loadcase))
        res_beam.merge_cells(
            'F%d:G%d' % (j * (num_loadcase + 6) + 5 + num_loadcase, j * (num_loadcase + 6) + 5 + num_loadcase))
        res_beam.merge_cells(
            'H%d:I%d' % (j * (num_loadcase + 6) + 5 + num_loadcase, j * (num_loadcase + 6) + 5 + num_loadcase))
        res_beam.merge_cells(
            'J%d:K%d' % (j * (num_loadcase + 6) + 5 + num_loadcase, j * (num_loadcase + 6) + 5 + num_loadcase))

    #输出结果
    # 分组指针，代表已输出的有结果的group数量
    group_pointer = 0
    #Excel列标
    word_bar = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    # group_stress指针，代表遍历所到的分组
    i = 0
    #绝对值最大的值
    minmax = 0
    while i < num_group:
        if group_stress[i][8][0]:
            # 组名
            res_beam['%s%d' % (
                word_bar[2 * (group_pointer % 5)], (int(group_pointer / 5) * (num_loadcase + 6) + 1))] = \
                name_group[i]
            for j in range(0, num_loadcase):
                #应力
                res_beam['%s%d' % (word_bar[2 * (group_pointer % 5)], (int(group_pointer / 5) * (num_loadcase + 6) + 3 + j))] = group_stress[i][9][j]
                res_beam['%s%d' % (word_bar[1 + 2 * (group_pointer % 5)], (int(group_pointer / 5) * (num_loadcase + 6) + 3 + j))] = group_stress[i][10][j]
                if max(abs(group_stress[i][9][j]), abs(group_stress[i][10][j])) > abs(minmax):
                    if abs(group_stress[i][9][j]) > abs(group_stress[i][10][j]):
                        minmax = group_stress[i][9][j]
                    else:
                        minmax = group_stress[i][10][j]
            res_beam['%s%d' % (word_bar[2 * (group_pointer % 5)], (int(group_pointer / 5) * (num_loadcase + 6) + 3 + num_loadcase))] = minmax
            minmax = 0
            i += 1
            group_pointer += 1
        else:
            i += 1
            # print(i)
            # print(447)

    #保存生成Excel文件
    res_excel.save(pcl_file_location + '/result.xlsx')


def debug():
    # global stress_standard
    # global stress_groupname
    # global stress_id
    # global stress_k
    #
    # print(stress_k)
    # print(stress_id)
    # print(stress_groupname)
    # print(stress_standard)
    read_rpt()

###################################################
#定义模块

l_pcl_location = tk.Label(root, text=pcl_file_location)
b_location = tk.Button(root, text = "目标文件夹", command = pcl_location)
b_ok = tk.Button(root, text = "生成PCL", command = create_pcl)
b_create_excel = tk.Button(root, text = "生成Excel文件", command = create_excel)

L_loadcase = tk.Label(root, text = "工况")
L_group = tk.Label(root, text = "分组")

loadcase_text = tk.Text(root, width=20, height=25)
group_text = tk.Text(root, width=20, height=25)
# group_text1 = tk.Text(root, width=20, height=15)

l_select_result = tk.Label(root, text = "选择结果文件形式")
b_op2 = tk.Radiobutton(root, text = "op2", variable = re_kind, value = 0)
b_xdb = tk.Radiobutton(root, text = "xdb, 编号：", variable = re_kind, value = 1)
e_xdb_num = tk.Entry(root)

b_debug = tk.Button(root, text = 'debug', command = debug)
##########################################################
#模块初始化

##########################################################
#显示模块
L_loadcase.grid(row = 3, column = 4)
L_group.grid(row = 3, column = 1)

loadcase_text.grid(row = 4, column = 4, rowspan = 10, columnspan = 2)
group_text.grid(row = 8, column = 1,columnspan = 2, rowspan = 4)
# group_text1.grid(row = 5, column = 1,columnspan = 2)
l_select_result.grid(row = 0, column = 1)
b_op2.grid(row = 0, column = 2)
b_xdb.grid(row = 0, column = 3)
e_xdb_num.grid(row = 0, column = 4)

b_location.grid(row = 1, column = 1)
b_ok.grid(row = 2, column = 1)
b_create_excel.grid(row = 2, column = 2)
# b_debug.grid(row = 2, column = 2)


root.mainloop()
